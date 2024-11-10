import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import sqlite3  # 添加 sqlite3 导入
import requests
from bs4 import BeautifulSoup
import time
import random
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill
import datetime
from tkinter import filedialog


def search_google(query, site, max_pages=50):
    query = f"{query}"
    for page in range(max_pages):
        url = f"https://www.google.com/search?q={query}&start={page * 10}"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
        }

        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            print(f"Failed to retrieve results: {response.status_code}")
            return False

        soup = BeautifulSoup(response.text, 'html.parser')
        results = soup.find('div', id='main').find_all('div', recursive=False)[3:13]   # 查找所有链接

        for index, result in enumerate(results):
            if site in result.text:
                return [page, index+1, page * 10 + index + 1, url]

        print(f"Page {page + 1} does not contain {site}. Continuing to next page...")
        time.sleep(random.randint(1, 3))

    print(f"{site} not found in the specified pages.")
    return None


class SiteKeywordManager:
    def __init__(self, root):
        self.root = root
        self.root.title("站点关键词Google排名查询器 v1.0")
        self.root.iconbitmap(".\\resource\\rank.ico")
        # 初始化数据库
        self.init_database()
        
        # 从数据库加载数据
        self.sites = []
        self.keywords = {}
        self.rank_update = False
        self.rank_error = False
        
        # 创建左右分栏
        self.left_frame = ttk.Frame(root)
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        
        self.right_frame = ttk.Frame(root)
        self.right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 左侧站点管理
        self.setup_site_section()
        
        # 右侧关键词管理
        self.setup_keyword_section()
        self.load_data()
        self.loop()

    def loop(self):
        if self.rank_update:
            selection = self.site_tree.selection()
            if selection:  # 如果没有选中项，直接返回
                selected_item = self.site_tree.item(selection[0])
                selected_site = selected_item['values'][1]
                self._update_database_ranks(selected_site, self.keywords[selected_site])
                print('排名更新到数据库')
                self.rank_update = False
                self.on_site_select(None)
                messagebox.showinfo("刷新排名", "关键词排名已刷新")

        self.root.after(500, self.loop)

    def init_database(self):
        """初始化数据库连接和表"""
        self.conn = sqlite3.connect('local.db')
        self.cursor = self.conn.cursor()
        
        # 创建站点表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS sites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE
            )
        ''')
        
        # 创建关键表，添加 rank 列
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS keywords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER,
                keyword TEXT,
                rank INTEGER DEFAULT 0,
                FOREIGN KEY (site_id) REFERENCES sites (id),
                UNIQUE(site_id, keyword)
            )
        ''')
        self.conn.commit()

    def load_data(self):
        """从数据库加载数据到存"""
        # 加载站点
        self.cursor.execute('SELECT name FROM sites order by id')
        sites = self.cursor.fetchall()
        self.sites = [site[0] for site in sites]
        
        # 加载每个站点的关键词和排名
        for index, site in enumerate(self.sites, 1):
            self.cursor.execute('''
                SELECT k.keyword, k.rank 
                FROM keywords k
                JOIN sites s ON k.site_id = s.id 
                WHERE s.name = ? order by k.id
            ''', (site,))
            self.keywords[site] = [(row[0], row[1]) for row in self.cursor.fetchall()]
            
            # 更新站点列表显示
            self.site_tree.insert('', 'end', values=(index, site))

    def setup_site_section(self):
        # 站点输入框
        site_frame = ttk.Frame(self.left_frame)
        site_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(site_frame, text="站点:").pack(side=tk.LEFT)
        self.site_entry = ttk.Entry(site_frame)
        self.site_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(site_frame, text="添加站点", command=self.add_site).pack(side=tk.LEFT)
        
        # 站点列表改为 Treeview
        ttk.Label(self.left_frame, text="站点列表:").pack(anchor=tk.W)
        columns = ('序号', '站点')
        self.site_tree = ttk.Treeview(self.left_frame, columns=columns, show='headings', height=15)
        
        # 置列标题和宽度
        self.site_tree.heading('序号', text='序号')
        self.site_tree.heading('站点', text='站点')
        self.site_tree.column('序号', width=50, anchor='center')
        self.site_tree.column('站点', width=150, anchor='w')
        
        self.site_tree.pack(fill=tk.BOTH, expand=True)
        self.site_tree.bind('<<TreeviewSelect>>', self.on_site_select)
        self.site_tree.bind('<Delete>', self.delete_site)

    def setup_keyword_section(self):
        # 关键词输入区域
        keyword_input_frame = ttk.Frame(self.right_frame)
        keyword_input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(keyword_input_frame, text="关键词:").pack(side=tk.LEFT)
        self.keyword_entry = ttk.Entry(keyword_input_frame)
        self.keyword_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(keyword_input_frame, text="添加关键词", 
                  command=self.add_keyword).pack(side=tk.LEFT)

        # 添加页码设置
        ttk.Label(keyword_input_frame, text="最大页码:").pack(side=tk.LEFT, padx=(10, 0))
        self.max_pages_entry = ttk.Entry(keyword_input_frame, width=5)
        self.max_pages_entry.insert(0, "50")  # 默认值为50
        self.max_pages_entry.pack(side=tk.LEFT, padx=5)

        self.refresh_button = ttk.Button(keyword_input_frame, text="刷新排名",
                                          command=self.refresh_ranks)
        self.refresh_button.pack(side=tk.LEFT, padx=5)

        # 添加进度条和进度标签到同一行
        progress_frame = ttk.Frame(self.right_frame)
        progress_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)  # 修改为填充底部
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.progress_label = ttk.Label(progress_frame, text="0/0")
        self.progress_label.pack(side=tk.LEFT, padx=5)

        # 添加导出按钮到关键词输入区域
        self.export_button = ttk.Button(keyword_input_frame, text="导出报告",
                                      command=self.export_report)
        self.export_button.pack(side=tk.LEFT, padx=5)

        # 关键词表格
        self.setup_keyword_table()

    def setup_keyword_table(self):
        # 站点列表
        ttk.Label(self.right_frame, text="关键词列表:").pack(anchor=tk.W)
        
        # 创建表格，添加排名列
        columns = ('序号', '关键词', '排名')
        self.keyword_tree = ttk.Treeview(self.right_frame, columns=columns, 
                                       show='headings')
        
        # 设置列标题
        self.keyword_tree.heading('序号', text='序号')
        self.keyword_tree.heading('关键词', text='关键词')
        self.keyword_tree.heading('排名', text='排名')
        
        # 设置列宽
        self.keyword_tree.column('序号', width=50, anchor='center')
        self.keyword_tree.column('关键词', width=200, anchor='w')
        self.keyword_tree.column('排名', width=100, anchor='center')
        
        # 直接示表格
        self.keyword_tree.pack(fill=tk.BOTH, expand=True)
        
        # 添加键盘绑定
        self.keyword_tree.bind('<Delete>', self.delete_keyword)

    def add_site(self):
        site = self.site_entry.get()
        
        if not site:
            messagebox.showerror("错误", "站点名称不能为空！")
            return
        
        try:
            # 添加站点到数据库
            self.cursor.execute('INSERT INTO sites (name) VALUES (?)', (site,))
            self.conn.commit()
            
            # 更新内存数据
            self.sites.append(site)
            self.keywords[site] = []
            
            # 更新界面
            self.site_tree.insert('', 'end', values=(len(self.sites), site))
            self.site_entry.delete(0, tk.END)
            
        except sqlite3.IntegrityError:
            messagebox.showerror("错误", "该站点存在！")
            return

    def on_site_select(self, event):
        # 获取选中的站点
        selection = self.site_tree.selection()
        if not selection:  # 如果没有选中项，直接返回
            return
            
        selected_item = self.site_tree.item(selection[0])
        selected_site = selected_item['values'][1]  # 获取站点名称
        
        # 清空当前表格
        for item in self.keyword_tree.get_children():
            self.keyword_tree.delete(item)
            
        # 重新设置表格的列
        self.keyword_tree.heading('序号', text='序号')
        self.keyword_tree.heading('关键词', text='关键词')
        self.keyword_tree.heading('排名', text='排名')
        
        # 设置列宽
        self.keyword_tree.column('序号', width=50, anchor='center')
        self.keyword_tree.column('关键词', width=200, anchor='w')
        self.keyword_tree.column('排名', width=100, anchor='center')
        
        # 填充该站点的关键词数据
        for index, (keyword, rank) in enumerate(self.keywords[selected_site], 1):
            display_rank = "无排名" if rank == 0 else rank
            self.keyword_tree.insert('', 'end', values=(index, keyword, display_rank))
            
        # 确保表格可见
        self.keyword_tree.pack(fill=tk.BOTH, expand=True)

    def add_keyword(self):
        # 获取当前选中的站点
        selection = self.site_tree.selection()
        if not selection:
            messagebox.showerror("错误", "请先选择一个站点！")
            return
            
        selected_item = self.site_tree.item(selection[0])
        selected_site = selected_item['values'][1]  # 获取站点名称
        keyword = self.keyword_entry.get().strip()
        
        if not keyword:
            messagebox.showerror("错误", "关键词不能为空！")
            return
            
        try:
            # 取站点ID
            self.cursor.execute('SELECT id FROM sites WHERE name = ?', (selected_site,))
            site_id = self.cursor.fetchone()[0]
            
            # 添加关键词到数据库，设置默认排名为0
            self.cursor.execute(''' 
                INSERT INTO keywords (site_id, keyword, rank) 
                VALUES (?, ?, 0)
            ''', (site_id, keyword))
            self.conn.commit()
            
            # 更新内存数据
            self.keywords[selected_site].append((keyword, 0))
            
            # 更新表格显示，显示“无排名”
            self.keyword_tree.insert('', 'end', 
                                   values=(len(self.keywords[selected_site]), keyword, "无排名"))  # 修改此行
            
            # 清空输入框
            self.keyword_entry.delete(0, tk.END)
            
        except sqlite3.IntegrityError:
            messagebox.showerror("错误", "该关键词已存在！")
            return

    def delete_site(self, event=None):
        """删除选中的站点"""
        selection = self.site_tree.selection()
        if not selection:
            messagebox.showerror("错误", "请先选择要删除的站点！")
            return
            
        selected_item = self.site_tree.item(selection[0])
        selected_site = selected_item['values'][1]  # 获取站点名称
        
        # 确认是否删除
        if not messagebox.askyesno("确认", f"确定要删除站点 '{selected_site}' 吗？\n这将同时删除该站点的所有关键词！"):
            return
            
        try:
            # 获取站点ID
            self.cursor.execute('SELECT id FROM sites WHERE name = ?', (selected_site,))
            site_id = self.cursor.fetchone()[0]
            
            # 删除关联的关键词
            self.cursor.execute('DELETE FROM keywords WHERE site_id = ?', (site_id,))
            
            # 删除站点
            self.cursor.execute('DELETE FROM sites WHERE id = ?', (site_id,))
            self.conn.commit()
            
            # 更新内存数据
            self.sites.remove(selected_site)
            del self.keywords[selected_site]
            
            # 更新界面
            self.site_tree.delete(selection[0])
            
            # 清空关键词表格
            for item in self.keyword_tree.get_children():
                self.keyword_tree.delete(item)
                
        except sqlite3.Error as e:
            messagebox.showerror("错误", f"删除站点失败：{str(e)}")

    def delete_keyword(self, event=None):
        """删除选中的关键词"""
        selection = self.keyword_tree.selection()
        if not selection:
            messagebox.showerror("错误", "请先选择要删除的关键词！")
            return
            
        # 获取当前选中的站点
        site_selection = self.site_tree.selection()
        if not site_selection:
            return
        selected_site = self.site_tree.item(site_selection[0])['values'][1]  # 获取站点名称
        
        # 获取选中的关键词
        selected_item = self.keyword_tree.item(selection[0])
        keyword = selected_item['values'][1]
        
        # 确认是否删除
        if not messagebox.askyesno("确认", f"确定要删除关键词 '{keyword}' 吗？"):
            return
            
        try:
            # 获取站点ID
            self.cursor.execute('SELECT id FROM sites WHERE name = ?', (selected_site,))
            site_id = self.cursor.fetchone()[0]
            
            # 从数据库删除关键词
            self.cursor.execute(''' 
                DELETE FROM keywords 
                WHERE site_id = ? AND keyword = ? 
            ''', (site_id, keyword))
            self.conn.commit()
            
            # 更新内存数据
            self.keywords[selected_site] = [kw for kw in self.keywords[selected_site] if kw[0] != keyword]  # 修改此行
            
            # 更新表格显示
            self.keyword_tree.delete(selection[0])
            
            # 重新排序显示
            self.on_site_select(None)
            
        except sqlite3.Error as e:
            messagebox.showerror("错误", f"删除关键词失败：{str(e)}")

    def refresh_ranks(self):
        """刷新选中站点的关键词排名"""
        selection = self.site_tree.selection()
        if not selection:
            messagebox.showerror("错误", "请先选择一个站点！")
            return
            
        selected_item = self.site_tree.item(selection[0])
        selected_site = selected_item['values'][1]
        
        # 获取当前站点的关键词
        updated_keywords = self.keywords[selected_site][:]  # 复制当前关键词

        # 获取最大页码设置
        try:
            max_pages = int(self.max_pages_entry.get())
        except ValueError:
            messagebox.showerror("错误", "请输入有效的页码！")
            return

        # 启动刷新排名的操作
        self._refresh_ranks_thread(selected_site, updated_keywords, max_pages)  # 传递 updated_keywords 和 max_pages

    def _refresh_ranks_thread(self, selected_site, updated_keywords, max_pages):
        self.rank_error = False
        # 禁用刷新按钮
        self.refresh_button.config(state=tk.DISABLED)
        """刷新关键词排名"""
        def worker(updated_keywords, event):
            # 更新内存数据结构
            total_keywords = len(updated_keywords)
            self.progress_bar['maximum'] = total_keywords
            self.progress_bar['value'] = 0
            
            for index, (keyword, _) in enumerate(updated_keywords):
                try:
                    new_rank_data = search_google(keyword, selected_site, max_pages)  # 使用 max_pages
                    rank_value = new_rank_data[2] if new_rank_data else 0
                    
                    # 更新关键词的排名
                    updated_keywords[index] = (keyword, rank_value)
                    
                except Exception as e:  # 捕捉网络连接异常
                    self.rank_error = True
                    break  # 结束线程
                
                # 更新进度条和标签
                self.progress_bar['value'] += 1
                self.progress_label['text'] = f"{self.progress_bar['value']}/{total_keywords}"  # 更新标签
                self.progress_bar.update_idletasks()
            
            # 在这里更新数据库中的排名
            self.keywords[selected_site] = updated_keywords
            event.set()  # 设置事件，表示任务完成

        event = threading.Event()  # 创建事件
        thread = threading.Thread(target=worker, args=(updated_keywords, event))
        thread.start()  # 启动线程

        # 监听线程完成
        thread_done = threading.Thread(target=self._check_thread_completion, args=(event,))
        thread_done.start()

    def _check_thread_completion(self, event):
        """检查线程是否完成"""
        event.wait()  # 等待事件被设置
        # 使刷新按钮可用
        self.refresh_button.config(state=tk.NORMAL)
        if self.rank_error:
            messagebox.showerror("错误", "获取排名时发生错误，请检查网络连接！")
        else:
            self.rank_update = True

    def _update_database_ranks(self, selected_site, updated_keywords):
        """更新数据库中的关键词排名"""
        # 获取站点ID
        self.cursor.execute('SELECT id FROM sites WHERE name = ?', (selected_site,))
        site_id = self.cursor.fetchone()[0]
        
        # 批量更新数据库中的排名
        for keyword, rank_value in updated_keywords:
            self.cursor.execute(''' 
                UPDATE keywords 
                SET rank = ? 
                WHERE site_id = ? AND keyword = ? 
            ''', (rank_value, site_id, keyword))
        
        self.conn.commit()  # 提交更改

    def stop_refresh(self):
        """停止刷新操作"""
        self.stop_refreshing = True

    def __del__(self):
        """析构函数，确保程序退出时关闭数据库连接"""
        if hasattr(self, 'conn'):
            self.conn.close()

    def export_report(self):
        """导出排名报告"""
        # 获取当前选中的站点
        selection = self.site_tree.selection()
        if not selection:
            messagebox.showerror("错误", "请先选择一个站点！")
            return
            
        selected_item = self.site_tree.item(selection[0])
        selected_site = selected_item['values'][1]
        
        # 获取当前时间作为文件名
        current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"{selected_site}_排名报告_{current_time}.xlsx"
        
        # 选择保存位置
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if not filename:
            return
            
        try:
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "排名报告"
            
            # 设置表头
            headers = ["序号", "关键词", "排名", "更新时间"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                
            # 设置表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # 写入数据
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for row, (keyword, rank) in enumerate(self.keywords[selected_site], 2):
                ws.cell(row=row, column=1, value=row-1)  # 序号
                ws.cell(row=row, column=2, value=keyword)  # 关键词
                ws.cell(row=row, column=3, value=rank if rank > 0 else "无排名")  # 排名
                ws.cell(row=row, column=4, value=current_time)  # 更新时间
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # 保存文件
            wb.save(filename)
            messagebox.showinfo("成功", "排名报告已导出！")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出报告失败：{str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = SiteKeywordManager(root)
    root.mainloop()

